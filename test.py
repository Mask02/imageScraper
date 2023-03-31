import tracemalloc
from asyncio import wait, FIRST_COMPLETED
from time import sleep
import requests
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
from concurrent.futures import ThreadPoolExecutor
from selenium.webdriver.chrome.options import Options
import openpyxl

tracemalloc.start()

workbook = openpyxl.load_workbook('data.xlsm')
sheet = workbook.active


def scrape_image(item):
    # options = uc.ChromeOptions()
    options = Options()
    options.add_argument("--headless")
    options.add_argument(
        '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36')
    # driver = uc.Chrome(options=options)
    driver = webdriver.Chrome()

    file_path = 'path.txt'
    with open(file_path, 'r') as f:
        path = f.read()

    input_path = rf'{path}'
    site_url = "https://www.astrobin.com/search/?q="

    complete_path = os.path.join(input_path, 'images')
    if not os.path.exists(complete_path):
        os.mkdir(complete_path)
    # print(item)
    driver.get(site_url)

    search_box = driver.find_element(By.XPATH, "//input[@id='q']")
    search_box.send_keys(item)
    sleep(1)
    search_box.send_keys(Keys.ENTER)
    sleep(5)
    image_urls = driver.find_elements(By.XPATH, "//figcaption/a")
    image_urls_array = []

    for urls in image_urls:
        image_urls_array.append(urls.get_attribute('href'))
    print(len(image_urls_array))

    for i, url in enumerate(image_urls_array):

        driver.get(url)
        while True:
            image_url = driver.find_element(By.XPATH, "//figure//img").get_attribute('src')
            if image_url.split('.')[-1] == "jpg" or image_url.split('.')[-1] == "png":
                break
            else:
                pass
        sleep(1)
        filename = os.path.join(complete_path, item + str(i + 1) + '.' + image_url.split('.')[-1])
        response = requests.get(image_url)
        if response.status_code == 200:
            with open(filename, "wb") as f:
                f.write(response.content)
            print(filename)
        else:
            print("Failed to download image")


MAX_THREADS = 8
MIN_THREADS = 2

with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
    futures = []
    for cell in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        for item in cell:
            futures.append(executor.submit(scrape_image, item))
        while len(futures) >= MAX_THREADS:
            done, futures = wait(futures, return_when=FIRST_COMPLETED)

