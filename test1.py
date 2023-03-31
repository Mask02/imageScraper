from time import sleep
import requests
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import asyncio
import openpyxl
import threading


workbook = openpyxl.load_workbook('data.xlsm')
sheet = workbook.active


async def scrape_image(item, semaphore):
    async with semaphore:
        options = Options()
        # options.add_argument("--headless")
        options.add_argument(
            '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36')
        driver = webdriver.Chrome(options=options)

        file_path = 'path.txt'
        with open(file_path, 'r') as f:
            path = f.read()

        input_path = rf'{path}'
        site_url = "https://www.astrobin.com/search/?q="

        complete_path = os.path.join(input_path, 'images')
        if not os.path.exists(complete_path):
            os.mkdir(complete_path)
        driver.get(site_url)

        search_box = driver.find_element(By.XPATH, "//input[@id='q']")
        search_box.send_keys(item)
        await asyncio.sleep(1)
        search_box.send_keys(Keys.ENTER)
        await asyncio.sleep(5)
        image_urls = driver.find_elements(By.XPATH, "//figcaption/a")
        image_urls_array = []

        for urls in image_urls:
            image_urls_array.append(urls.get_attribute('href'))
        print(len(image_urls_array))

        for i, url in enumerate(image_urls_array):

            driver.get(url)
            while True:
                image_url = driver.find_element(By.XPATH, "//figure//img").get_attribute('src')
                if image_url.split('.')[-1] == "jpg" or image_url.split('.')[-1] == "png":
                    break
                else:
                    pass
            await asyncio.sleep(1)
            filename = os.path.join(complete_path, item + str(i + 1) + '.jpg')
            response = requests.get(image_url)
            if response.status_code == 200:
                with open(filename, "wb") as f:
                    f.write(response.content)
                print(f"Image saved successfully!{i + 1}")
            else:
                print("Failed to download image")

        driver.quit()


async def main():
    MAX_CONCURRENT_THREADS = 5
    MIN_CONCURRENT_THREADS = 2

    semaphore = asyncio.Semaphore(MAX_CONCURRENT_THREADS)
    tasks = []
    for cell in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        for item in cell:
            task = asyncio.create_task(scrape_image(item, semaphore))
            tasks.append(task)
            while len(tasks) >= MAX_CONCURRENT_THREADS:
                done, _ = await asyncio.wait(tasks, return_when=asyncio.FIRST_COMPLETED)
                tasks = [t for t in tasks if not t.done()]

    if tasks:
        await asyncio.wait(tasks)


if __name__ == "__main__":
    asyncio.run(main())
